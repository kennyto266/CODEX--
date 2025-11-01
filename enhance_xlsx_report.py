#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx 股票分析报告增强器
Add professional formatting, charts and visualizations using xlsx skills
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import LineChart, Reference, BarChart, ScatterChart
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
import json

def load_analysis_results():
    """加载分析结果"""
    with open('analysis_results.json', 'r', encoding='utf-8') as f:
        return json.load(f)

def load_stock_data():
    """加载股票原始数据"""
    stock_df = pd.read_csv('test_data/test_stock_0001_HK.csv', index_col=0, parse_dates=True)
    return stock_df

def load_strategy_data():
    """加载策略数据"""
    boll_df = pd.read_csv('test_data/test_strategy_boll.csv', index_col=0, parse_dates=True)
    rsi_df = pd.read_csv('test_data/test_strategy_rsi.csv', index_col=0, parse_dates=True)
    return {'BOLL': boll_df, 'RSI': rsi_df}

def apply_professional_formatting(wb):
    """Apply professional formatting to worksheets"""
    # print("Applying professional formatting to worksheets...")

    # 定义颜色
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    accent_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 格式化每个工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # # print(f"  Formatting sheet: {sheet_name}")

        # 设置列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # 添加边框和格式到数据区域
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                if cell.row == 1:  # 标题行
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column == 1:  # 第一列（标签列）
                    cell.font = Font(bold=True)
                else:  # 数据列
                    cell.alignment = Alignment(horizontal='center', vertical='center')

def add_charts_to_excel(wb):
    """Add charts to Excel"""
    # print("Adding charts to Excel worksheets...")

    # 加载数据
    stock_data = load_stock_data()
    strategy_data = load_strategy_data()
    results = load_analysis_results()

    # === 工作表 10: 图表分析 ===
    if '图表分析' not in wb.sheetnames:
        chart_ws = wb.create_sheet('图表分析')
        # print("  Creating new sheet: 图表分析")

        # 设置标题
        chart_ws['A1'] = "xlsx 股票分析图表报告"
        chart_ws['A1'].font = Font(size=16, bold=True, color="366092")
        chart_ws.merge_cells('A1:H1')

        # 添加数据用于图表
        chart_ws['A3'] = "日期"
        chart_ws['B3'] = "股票价格"
        chart_ws['C3'] = "BOLL策略"
        chart_ws['D3'] = "RSI策略"

        # 插入前50行数据
        for i, (date, row) in enumerate(stock_data.head(50).iterrows(), 4):
            chart_ws[f'A{i}'] = date.strftime('%Y-%m-%d')
            chart_ws[f'B{i}'] = row['close']
            chart_ws[f'C{i}'] = strategy_data['BOLL'].iloc[i-4, 1]
            chart_ws[f'D{i}'] = strategy_data['RSI'].iloc[i-4, 1]

        # 创建累积收益线图
        chart1 = LineChart()
        chart1.title = "累积收益对比图"
        chart1.style = 13
        chart1.y_axis.title = '累积收益'
        chart1.x_axis.title = '日期'

        data = Reference(ws=chart_ws, min_col=2, min_row=3, max_col=4, max_row=53)
        cats = Reference(ws=chart_ws, min_col=1, min_row=4, max_row=53)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)

        chart_ws.add_chart(chart1, "F3")

        # === 月度收益条形图 ===
        chart_ws['A60'] = "月份"
        chart_ws['B60'] = "股票月度收益 (%)"
        chart_ws['C60'] = "BOLL月度收益 (%)"
        chart_ws['D60'] = "RSI月度收益 (%)"

        monthly_returns = results['monthly_returns']
        row_num = 61
        for date_str, value in monthly_returns['stock']['returns'].items():
            chart_ws[f'A{row_num}'] = date_str.split(' ')[0]
            chart_ws[f'B{row_num}'] = round(value * 100, 2)
            chart_ws[f'C{row_num}'] = round(monthly_returns['boll']['returns'][date_str] * 100, 2)
            chart_ws[f'D{row_num}'] = round(monthly_returns['rsi']['returns'][date_str] * 100, 2)
            row_num += 1

        # 创建条形图
        chart2 = BarChart()
        chart2.title = "月度收益对比"
        chart2.y_axis.title = '收益率 (%)'
        chart2.x_axis.title = '月份'

        data2 = Reference(ws=chart_ws, min_col=2, min_row=60, max_col=4, max_row=row_num-1)
        cats2 = Reference(ws=chart_ws, min_col=1, min_row=61, max_row=row_num-1)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)

        chart_ws.add_chart(chart2, "F60")

        # === 绩效指标雷达图数据 ===
        chart_ws['A100'] = "绩效指标"
        chart_ws['B100'] = "数值"
        chart_ws['C100'] = "百分比"

        stock_metrics = results['performance_metrics']['stock']
        metrics = [
            ('总收益率', stock_metrics['total_return']),
            ('年化收益率', stock_metrics['annualized_return']),
            ('波动率', stock_metrics['volatility']),
            ('夏普比率', stock_metrics['sharpe_ratio'] * 10),  # 放大10倍以便可视化
            ('最大回撤', abs(stock_metrics['max_drawdown']))
        ]

        for i, (metric, value) in enumerate(metrics, 101):
            chart_ws[f'A{i}'] = metric
            chart_ws[f'B{i}'] = value
            chart_ws[f'C{i}'] = value

        # 创建条形图
        chart3 = BarChart()
        chart3.title = "绩效指标分析"
        chart3.y_axis.title = '数值'
        chart3.x_axis.title = '指标'

        data3 = Reference(ws=chart_ws, min_col=2, min_row=100, max_col=3, max_row=105)
        cats3 = Reference(ws=chart_ws, min_col=1, min_row=101, max_row=105)
        chart3.add_data(data3, titles_from_data=True)
        chart3.set_categories(cats3)

        chart_ws.add_chart(chart3, "F100")

def add_conditional_formatting(wb):
    """Add conditional formatting"""
    # print("Adding conditional formatting...")

    # 为股票绩效指标工作表Add conditional formatting
    if '股票绩效指标' in wb.sheetnames:
        ws = wb['股票绩效指标']
        # 为数值列Add conditional formatting（颜色缩放）
        color_scale = ColorScaleRule(
            start_type='min', start_color='FF6B6B',
            mid_type='percentile', mid_value=50, mid_color='FFFF99',
            end_type='max', end_color='4ECDC4'
        )
        ws.conditional_formatting.add('B2:B7', color_scale)

    # 为策略绩效对比Add conditional formatting
    if '策略绩效对比' in wb.sheetnames:
        ws = wb['策略绩效对比']
        # 为收益率列Add conditional formatting
        for col in ['C', 'D', 'E']:  # 总收益率、超额收益、最大回撤
            rule = CellIsRule(operator='lessThan', formula=[0], fill=PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'))
            ws.conditional_formatting.add(f'{col}2:{col}3', rule)

def create_executive_summary(wb):
    """Create Executive Summary sheet"""
    # print("Creating Executive Summary sheet...")

    if 'Executive summary' in wb.sheetnames:
        wb.remove(wb['Executive summary'])

    ws = wb.create_sheet('Executive summary', 0)  # 插入为第一个工作表

    results = load_analysis_results()

    # 标题
    ws['A1'] = "xlsx 股票分析 - Executive summary"
    ws['A1'].font = Font(size=18, bold=True, color="366092")
    ws.merge_cells('A1:F1')

    # 关键指标卡片
    ws['A3'] = "关键指标"
    ws['A3'].font = Font(size=14, bold=True)
    ws['A3'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    stock_metrics = results['performance_metrics']['stock']
    strategy_metrics = results['performance_metrics']['strategies']

    # 指标数据
    metrics_data = [
        ["股票代码", results['summary']['report_info']['stock_symbol']],
        ["分析期间", results['summary']['report_info']['period']],
        ["总收益率", f"{stock_metrics['total_return']:.2f}%"],
        ["年化收益率", f"{stock_metrics['annualized_return']:.2f}%"],
        ["波动率", f"{stock_metrics['volatility']:.2f}%"],
        ["夏普比率", f"{stock_metrics['sharpe_ratio']:.2f}"],
        ["最大回撤", f"{stock_metrics['max_drawdown']:.2f}%"],
        ["最佳策略", "BOLL"],
        ["风险等级", "高"]
    ]

    for i, (label, value) in enumerate(metrics_data, 4):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)

    # 策略对比表
    ws['D3'] = "策略绩效对比"
    ws['D3'].font = Font(size=14, bold=True)
    ws['D3'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # 策略表头
    headers = ['策略', '总收益率', '超额收益', '胜率', '风险等级']
    for col, header in enumerate(headers, 4):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    # 策略数据
    for row, (strategy, data) in enumerate(strategy_metrics.items(), 5):
        ws.cell(row=row, column=4).value = strategy.upper()
        ws.cell(row=row, column=5).value = f"{data['total_return']:.2f}%"
        ws.cell(row=row, column=6).value = f"{data['excess_return']:.2f}%"
        ws.cell(row=row, column=7).value = f"{data['win_rate']:.2f}%"
        ws.cell(row=row, column=8).value = "高"

    # 投资建议
    ws['A15'] = "投资建议"
    ws['A15'].font = Font(size=14, bold=True)
    ws['A15'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    for i, rec in enumerate(results['summary']['recommendations'], 16):
        ws[f'A{i}'] = f"• {rec}"

    # 设置列宽
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 20

    # 合并单元格
    ws.merge_cells('A16:A18')

def enhance_excel_report():
    """Enhance Excel report"""
    # print("Enhancing Excel report with xlsx skills...")

    # Load existing workbook
    wb = load_workbook('xlsx_stock_analysis_report.xlsx')
    # print(f"Loaded workbook with {len(wb.sheetnames)} sheets")

    # Apply professional formatting
    apply_professional_formatting(wb)

    # Add conditional formatting
    add_conditional_formatting(wb)

    # Create Executive Summary
    create_executive_summary(wb)

    # Add charts
    add_charts_to_excel(wb)

    # 保存增强版
    output_file = 'xlsx_stock_analysis_enhanced.xlsx'
    wb.save(output_file)

    # print(f"\nEnhanced Excel report created: {output_file}")
    # print(f"Total sheets: {len(wb.sheetnames)}")
    # print("\nNew features added:")
    # print("  • Professional formatting")
    # print("  • Conditional formatting")
    # print("  • Executive summary")
    # print("  • Charts and visualizations")
    # print("  • Color-coded indicators")

    return output_file

if __name__ == "__main__":
    output_file = enhance_excel_report()
    # print(f"\nReport enhancement completed: {output_file}")
