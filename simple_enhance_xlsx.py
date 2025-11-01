#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsx Excel Report Enhancer - Simple Version
Add professional formatting and conditional formatting
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

def enhance_excel():
    """Enhance Excel report with formatting"""
    # Load workbook
    wb = load_workbook('xlsx_stock_analysis_report.xlsx')

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply formatting to each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Set column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Add borders and formatting
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                if cell.row == 1:  # Header row
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column == 1:  # First column
                    cell.font = Font(bold=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add conditional formatting to Stock Performance sheet
    if '股票绩效指标' in wb.sheetnames:
        ws = wb['股票绩效指标']
        color_scale = ColorScaleRule(
            start_type='min', start_color='FF6B6B',
            mid_type='percentile', mid_value=50, mid_color='FFFF99',
            end_type='max', end_color='4ECDC4'
        )
        ws.conditional_formatting.add('B2:B7', color_scale)

    # Add conditional formatting to Strategy Performance sheet
    if '策略绩效对比' in wb.sheetnames:
        ws = wb['策略绩效对比']
        for col in ['C', 'D', 'E']:
            rule = CellIsRule(operator='lessThan', formula=[0],
                            fill=PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid'))
            ws.conditional_formatting.add(f'{col}2:{col}3', rule)

    # Create Executive Summary sheet
    if '执行摘要' in wb.sheetnames:
        wb.remove(wb['执行摘要'])

    ws = wb.create_sheet('执行摘要', 0)

    # Title
    ws['A1'] = "xlsx Stock Analysis - Executive Summary"
    ws['A1'].font = Font(size=18, bold=True, color="366092")
    ws.merge_cells('A1:F1')

    # Key metrics
    ws['A3'] = "Key Metrics"
    ws['A3'].font = Font(size=14, bold=True)
    ws['A3'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    metrics_data = [
        ["Stock Symbol", "0001.HK"],
        ["Analysis Period", "2020-01-02 to 2021-01-07"],
        ["Total Return", "-23.49%"],
        ["Annualized Return", "-23.57%"],
        ["Volatility", "33.19%"],
        ["Sharpe Ratio", "-0.71"],
        ["Max Drawdown", "-39.51%"],
        ["Best Strategy", "BOLL"],
        ["Risk Level", "High"]
    ]

    for i, (label, value) in enumerate(metrics_data, 4):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)

    # Strategy comparison
    ws['D3'] = "Strategy Performance"
    ws['D3'].font = Font(size=14, bold=True)
    ws['D3'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    headers = ['Strategy', 'Total Return', 'Excess Return', 'Win Rate', 'Risk Level']
    for col, header in enumerate(headers, 4):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    # Strategy data
    strategy_data = [
        ["BOLL", "-23.49%", "-1.10%", "44.05%", "High"],
        ["RSI", "-23.49%", "-23.49%", "44.05%", "High"]
    ]

    for row, data in enumerate(strategy_data, 5):
        for col, value in enumerate(data, 4):
            ws.cell(row=row, column=col).value = value

    # Investment recommendations
    ws['A15'] = "Investment Recommendations"
    ws['A15'].font = Font(size=14, bold=True)
    ws['A15'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    recommendations = [
        "• Best performing strategy: BOLL",
        "• Consider using technical strategies to improve risk-adjusted returns",
        "• High volatility detected - consider position sizing and stop-loss strategies"
    ]

    for i, rec in enumerate(recommendations, 16):
        ws[f'A{i}'] = rec

    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 20

    # Save enhanced workbook
    output_file = 'xlsx_stock_analysis_enhanced.xlsx'
    wb.save(output_file)

    return output_file

if __name__ == "__main__":
    output = enhance_excel()
    print(f"Enhanced Excel report created: {output}")
